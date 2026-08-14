import streamlit as st
import pandas as pd

from io import BytesIO
from datetime import date

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =========================================================
# CONFIGURAÇÃO DA PÁGINA
# =========================================================

st.set_page_config(
    page_title="RH People Analytics",
    page_icon="📊",
    layout="wide"
)

st.title("📊 RH People Analytics")

st.write(
    """
    Dashboard gerencial para análise de indicadores,
    movimentações, estrutura e custos de Recursos Humanos.
    """
)

st.divider()


# =========================================================
# FUNÇÕES AUXILIARES
# =========================================================

def moeda(valor):
    """
    Converte número para formato monetário brasileiro.
    """

    if pd.isna(valor):
        valor = 0

    texto = f"{float(valor):,.2f}"

    texto = (
        texto
        .replace(",", "X")
        .replace(".", ",")
        .replace("X", ".")
    )

    return f"R$ {texto}"


def percentual(valor):
    """
    Converte número para percentual.
    """

    if pd.isna(valor):
        valor = 0

    return f"{float(valor):.2f}%".replace(".", ",")


def empregado_ativo_na_data(linha, data_referencia):
    """
    Verifica se o colaborador estava ativo
    em uma determinada data.
    """

    admissao = linha["Admissão"]
    demissao = linha["Demissão"]

    if pd.isna(admissao):
        return False

    if admissao > data_referencia:
        return False

    if pd.isna(demissao):
        return True

    return demissao >= data_referencia


def calcular_tempo_empresa_anos(data_admissao):
    """
    Calcula o tempo aproximado de empresa em anos.
    """

    if pd.isna(data_admissao):
        return 0

    hoje = pd.Timestamp(date.today())

    dias = (
        hoje
        - data_admissao
    ).days

    return max(
        dias / 365.25,
        0
    )


# =========================================================
# GERAÇÃO DA PLANILHA MODELO
# =========================================================

def gerar_planilha_modelo():

    dados_modelo = pd.DataFrame(
        {
            "Matrícula": [
                1001, 1002, 1003, 1004,
                1005, 1006, 1007, 1008,
                1009, 1010, 1011, 1012,
                1013, 1014, 1015, 1016,
                1017, 1018
            ],

            "Nome": [
                "João Silva",
                "Maria Souza",
                "Carlos Santos",
                "Ana Oliveira",
                "Pedro Lima",
                "Juliana Costa",
                "Ricardo Alves",
                "Fernanda Rocha",
                "Marcos Pereira",
                "Camila Santos",
                "Lucas Oliveira",
                "Patrícia Lima",
                "Rafael Costa",
                "Beatriz Alves",
                "Daniel Rocha",
                "Larissa Santos",
                "Paulo Fernandes",
                "Renata Lima"
            ],

            "Admissão": [
                "10/01/2024",
                "15/03/2023",
                "01/06/2022",
                "10/02/2025",
                "20/05/2024",
                "03/08/2021",
                "15/01/2026",
                "02/09/2023",
                "10/02/2026",
                "15/04/2026",
                "01/07/2025",
                "20/03/2024",
                "10/03/2026",
                "05/05/2026",
                "15/06/2026",
                "05/07/2026",
                "20/01/2025",
                "01/08/2026"
            ],

            "Demissão": [
                "",
                "10/08/2026",
                "",
                "",
                "",
                "",
                "",
                "30/06/2026",
                "",
                "",
                "15/05/2026",
                "",
                "",
                "",
                "20/07/2026",
                "",
                "",
                ""
            ],

            "Cargo": [
                "Analista de RH",
                "Assistente de DP",
                "Coordenador de RH",
                "Analista Financeiro",
                "Técnico de Segurança",
                "Analista de RH",
                "Assistente Administrativo",
                "Analista de DP",
                "Analista de Sistemas",
                "Assistente Financeiro",
                "Assistente Administrativo",
                "Analista de RH",
                "Desenvolvedor",
                "Assistente de RH",
                "Assistente Financeiro",
                "Analista de DP",
                "Técnico de Segurança",
                "Analista Financeiro"
            ],

            "Setor": [
                "RH",
                "DP",
                "RH",
                "Financeiro",
                "Segurança",
                "RH",
                "Administrativo",
                "DP",
                "TI",
                "Financeiro",
                "Administrativo",
                "RH",
                "TI",
                "RH",
                "Financeiro",
                "DP",
                "Segurança",
                "Financeiro"
            ],

            "Unidade": [
                "Matriz",
                "Matriz",
                "Matriz",
                "Filial Recife",
                "Obra Natal",
                "Matriz",
                "Filial Recife",
                "Matriz",
                "Matriz",
                "Filial Recife",
                "Obra Natal",
                "Matriz",
                "Matriz",
                "Filial Recife",
                "Filial Recife",
                "Matriz",
                "Obra Natal",
                "Matriz"
            ],

            "Salário": [
                4500,
                3200,
                8500,
                4800,
                5200,
                5800,
                2800,
                4700,
                6500,
                3500,
                2900,
                5100,
                7200,
                3000,
                3600,
                4900,
                5400,
                5200
            ],

            "Status": [
                "Ativo",
                "Desligado",
                "Ativo",
                "Ativo",
                "Ativo",
                "Ativo",
                "Ativo",
                "Desligado",
                "Ativo",
                "Ativo",
                "Desligado",
                "Ativo",
                "Ativo",
                "Ativo",
                "Desligado",
                "Ativo",
                "Ativo",
                "Ativo"
            ]
        }
    )

    buffer = BytesIO()

    with pd.ExcelWriter(
        buffer,
        engine="openpyxl"
    ) as writer:

        dados_modelo.to_excel(
            writer,
            index=False,
            sheet_name="Colaboradores"
        )

    return buffer.getvalue()


# =========================================================
# EXPORTAÇÃO PROFISSIONAL PARA EXCEL
# =========================================================

def gerar_relatorio_excel(
    dados_filtrados,
    resumo_setor,
    resumo_unidade,
    movimentacoes,
    indicadores
):

    buffer = BytesIO()

    with pd.ExcelWriter(
        buffer,
        engine="openpyxl"
    ) as writer:

        # -------------------------------------------------
        # ABA BASE
        # -------------------------------------------------

        dados_filtrados.to_excel(
            writer,
            sheet_name="Base Filtrada",
            index=False
        )

        # -------------------------------------------------
        # ABA SETORES
        # -------------------------------------------------

        resumo_setor.to_excel(
            writer,
            sheet_name="Resumo Setores",
            index=False
        )

        # -------------------------------------------------
        # ABA UNIDADES
        # -------------------------------------------------

        resumo_unidade.to_excel(
            writer,
            sheet_name="Resumo Unidades",
            index=False
        )

        # -------------------------------------------------
        # ABA MOVIMENTAÇÕES
        # -------------------------------------------------

        movimentacoes.to_excel(
            writer,
            sheet_name="Movimentacoes",
            index=False
        )

        # -------------------------------------------------
        # ABA INDICADORES
        # -------------------------------------------------

        pd.DataFrame(
            indicadores.items(),
            columns=[
                "Indicador",
                "Valor"
            ]
        ).to_excel(
            writer,
            sheet_name="Indicadores",
            index=False
        )

        # =================================================
        # FORMATAÇÃO
        # =================================================

        cabecalho_fill = PatternFill(
            "solid",
            fgColor="1F4E78"
        )

        cabecalho_font = Font(
            bold=True,
            color="FFFFFF"
        )

        borda = Border(
            bottom=Side(
                style="thin",
                color="D9D9D9"
            )
        )

        for nome_aba in writer.book.sheetnames:

            planilha = writer.book[nome_aba]

            planilha.freeze_panes = "A2"

            planilha.auto_filter.ref = (
                planilha.dimensions
            )

            # ---------------------------------------------
            # CABEÇALHO
            # ---------------------------------------------

            for celula in planilha[1]:

                celula.fill = cabecalho_fill
                celula.font = cabecalho_font

                celula.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            # ---------------------------------------------
            # LARGURA DAS COLUNAS
            # ---------------------------------------------

            for coluna in planilha.columns:

                maior = 0

                letra = get_column_letter(
                    coluna[0].column
                )

                for celula in coluna:

                    if celula.value is not None:

                        tamanho = len(
                            str(celula.value)
                        )

                        maior = max(
                            maior,
                            tamanho
                        )

                    celula.border = borda

                planilha.column_dimensions[
                    letra
                ].width = min(
                    maior + 3,
                    40
                )

    return buffer.getvalue()


# =========================================================
# DOWNLOAD DA PLANILHA MODELO
# =========================================================

st.subheader(
    "1. Modelo da base de colaboradores"
)

st.write(
    """
    Baixe o modelo para conhecer a estrutura
    esperada pelo sistema.
    """
)

modelo_excel = gerar_planilha_modelo()

st.download_button(
    label="📥 Baixar planilha modelo",
    data=modelo_excel,
    file_name="modelo_people_analytics.xlsx",
    mime=(
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    )
)

st.divider()


# =========================================================
# IMPORTAÇÃO
# =========================================================

st.subheader(
    "2. Importar base de colaboradores"
)

arquivo = st.file_uploader(
    "Selecione um arquivo Excel ou CSV",
    type=[
        "xlsx",
        "csv"
    ]
)


if arquivo is not None:

    # =====================================================
    # LEITURA
    # =====================================================

    try:

        if arquivo.name.lower().endswith(
            ".xlsx"
        ):

            dados = pd.read_excel(
                arquivo
            )

        else:

            dados = pd.read_csv(
                arquivo
            )

        st.success(
            "Base carregada com sucesso!"
        )

    except Exception as erro:

        st.error(
            f"Erro ao carregar arquivo: {erro}"
        )

        st.stop()


    # =====================================================
    # VALIDAÇÃO
    # =====================================================

    colunas_obrigatorias = [
        "Matrícula",
        "Nome",
        "Admissão",
        "Demissão",
        "Cargo",
        "Setor",
        "Unidade",
        "Salário",
        "Status"
    ]

    faltantes = [
        coluna
        for coluna in colunas_obrigatorias
        if coluna not in dados.columns
    ]

    if faltantes:

        st.error(
            "Existem colunas obrigatórias ausentes."
        )

        st.write(
            faltantes
        )

        st.stop()


    # =====================================================
    # TRATAMENTO
    # =====================================================

    dados["Admissão"] = pd.to_datetime(
        dados["Admissão"],
        dayfirst=True,
        errors="coerce"
    )

    dados["Demissão"] = pd.to_datetime(
        dados["Demissão"],
        dayfirst=True,
        errors="coerce"
    )

    dados["Salário"] = pd.to_numeric(
        dados["Salário"],
        errors="coerce"
    ).fillna(0)

    dados["Status"] = (
        dados["Status"]
        .astype(str)
        .str.strip()
        .str.title()
    )

    dados["TempoEmpresaAnos"] = (
        dados["Admissão"]
        .apply(
            calcular_tempo_empresa_anos
        )
    )


    # =====================================================
    # FILTROS
    # =====================================================

    st.sidebar.header(
        "🔎 Filtros"
    )

    setores = sorted(
        dados["Setor"]
        .dropna()
        .unique()
        .tolist()
    )

    unidades = sorted(
        dados["Unidade"]
        .dropna()
        .unique()
        .tolist()
    )

    cargos = sorted(
        dados["Cargo"]
        .dropna()
        .unique()
        .tolist()
    )


    filtro_setor = (
        st.sidebar.multiselect(
            "Setor",
            setores
        )
    )

    filtro_unidade = (
        st.sidebar.multiselect(
            "Unidade",
            unidades
        )
    )

    filtro_cargo = (
        st.sidebar.multiselect(
            "Cargo",
            cargos
        )
    )


    # =====================================================
    # ANO
    # =====================================================

    anos = sorted(
        set(
            dados["Admissão"]
            .dropna()
            .dt.year
            .tolist()
            +
            dados["Demissão"]
            .dropna()
            .dt.year
            .tolist()
            +
            [date.today().year]
        ),
        reverse=True
    )

    ano_analise = (
        st.sidebar.selectbox(
            "Ano de análise",
            anos
        )
    )


    # =====================================================
    # APLICAÇÃO DOS FILTROS
    # =====================================================

    dados_filtrados = dados.copy()

    if filtro_setor:

        dados_filtrados = (
            dados_filtrados[
                dados_filtrados[
                    "Setor"
                ].isin(
                    filtro_setor
                )
            ]
        )

    if filtro_unidade:

        dados_filtrados = (
            dados_filtrados[
                dados_filtrados[
                    "Unidade"
                ].isin(
                    filtro_unidade
                )
            ]
        )

    if filtro_cargo:

        dados_filtrados = (
            dados_filtrados[
                dados_filtrados[
                    "Cargo"
                ].isin(
                    filtro_cargo
                )
            ]
        )


    # =====================================================
    # ATIVOS / DESLIGADOS
    # =====================================================

    ativos = (
        dados_filtrados[
            dados_filtrados[
                "Status"
            ] == "Ativo"
        ]
        .copy()
    )

    desligados = (
        dados_filtrados[
            dados_filtrados[
                "Status"
            ] == "Desligado"
        ]
        .copy()
    )


    # =====================================================
    # PERÍODO
    # =====================================================

    inicio_ano = pd.Timestamp(
        year=ano_analise,
        month=1,
        day=1
    )

    if ano_analise == date.today().year:

        fim_ano = pd.Timestamp(
            date.today()
        )

    else:

        fim_ano = pd.Timestamp(
            year=ano_analise,
            month=12,
            day=31
        )


    # =====================================================
    # MOVIMENTAÇÕES
    # =====================================================

    admissoes_ano = (
        dados_filtrados[
            dados_filtrados[
                "Admissão"
            ].between(
                inicio_ano,
                fim_ano
            )
        ]
    )

    desligamentos_ano = (
        dados_filtrados[
            dados_filtrados[
                "Demissão"
            ].between(
                inicio_ano,
                fim_ano
            )
        ]
    )

    qtd_admissoes = len(
        admissoes_ano
    )

    qtd_desligamentos = len(
        desligamentos_ano
    )


    # =====================================================
    # HEADCOUNT MÉDIO
    # =====================================================

    hc_inicio = (
        dados_filtrados.apply(
            empregado_ativo_na_data,
            axis=1,
            data_referencia=inicio_ano
        )
        .sum()
    )

    hc_fim = (
        dados_filtrados.apply(
            empregado_ativo_na_data,
            axis=1,
            data_referencia=fim_ano
        )
        .sum()
    )

    headcount_medio = (
        hc_inicio
        + hc_fim
    ) / 2


    # =====================================================
    # TURNOVER
    # =====================================================

    if headcount_medio > 0:

        turnover = (
            qtd_desligamentos
            / headcount_medio
        ) * 100

    else:

        turnover = 0


    # =====================================================
    # INDICADORES
    # =====================================================

    headcount = len(
        ativos
    )

    folha = (
        ativos[
            "Salário"
        ].sum()
    )

    salario_medio = (
        ativos[
            "Salário"
        ].mean()
        if not ativos.empty
        else 0
    )

    maior_salario = (
        ativos[
            "Salário"
        ].max()
        if not ativos.empty
        else 0
    )

    menor_salario = (
        ativos[
            "Salário"
        ].min()
        if not ativos.empty
        else 0
    )

    tempo_medio = (
        ativos[
            "TempoEmpresaAnos"
        ].mean()
        if not ativos.empty
        else 0
    )


    # =====================================================
    # DASHBOARD EXECUTIVO
    # =====================================================

    st.subheader(
        "📊 Dashboard Executivo"
    )

    c1, c2, c3 = st.columns(3)

    c1.metric(
        "Headcount ativo",
        headcount
    )

    c2.metric(
        "Admissões no ano",
        qtd_admissoes
    )

    c3.metric(
        "Desligamentos no ano",
        qtd_desligamentos
    )


    c4, c5, c6 = st.columns(3)

    c4.metric(
        "Folha salarial mensal",
        moeda(
            folha
        )
    )

    c5.metric(
        "Salário médio",
        moeda(
            salario_medio
        )
    )

    c6.metric(
        "Turnover",
        percentual(
            turnover
        )
    )


    c7, c8, c9 = st.columns(3)

    c7.metric(
        "Tempo médio de empresa",
        f"{tempo_medio:.1f} anos"
    )

    c8.metric(
        "Maior salário",
        moeda(
            maior_salario
        )
    )

    c9.metric(
        "Menor salário",
        moeda(
            menor_salario
        )
    )


    st.caption(
        f"Ano analisado: {ano_analise}"
    )

    st.caption(
        "Turnover = desligamentos do período "
        "÷ headcount médio."
    )

    st.divider()


    # =====================================================
    # MOVIMENTAÇÃO MENSAL
    # =====================================================

    meses = [
        "Jan",
        "Fev",
        "Mar",
        "Abr",
        "Mai",
        "Jun",
        "Jul",
        "Ago",
        "Set",
        "Out",
        "Nov",
        "Dez"
    ]

    movimentacoes = pd.DataFrame(
        {
            "Mês": meses,
            "Admissões": [0] * 12,
            "Desligamentos": [0] * 12
        }
    )

    for mes in range(
        1,
        13
    ):

        movimentacoes.loc[
            mes - 1,
            "Admissões"
        ] = (
            admissoes_ano[
                admissoes_ano[
                    "Admissão"
                ].dt.month == mes
            ]
            .shape[0]
        )

        movimentacoes.loc[
            mes - 1,
            "Desligamentos"
        ] = (
            desligamentos_ano[
                desligamentos_ano[
                    "Demissão"
                ].dt.month == mes
            ]
            .shape[0]
        )


    st.subheader(
        "📈 Admissões x Desligamentos por mês"
    )

    st.bar_chart(
        movimentacoes.set_index(
            "Mês"
        )[
            [
                "Admissões",
                "Desligamentos"
            ]
        ]
    )


    # =====================================================
    # HEADCOUNT POR SETOR
    # =====================================================

    st.subheader(
        "👥 Headcount por setor"
    )

    headcount_setor = (
        ativos
        .groupby(
            "Setor"
        )
        .size()
        .sort_values(
            ascending=False
        )
    )

    if not headcount_setor.empty:

        st.bar_chart(
            headcount_setor
        )


    # =====================================================
    # HEADCOUNT POR CARGO
    # =====================================================

    st.subheader(
        "💼 Headcount por cargo"
    )

    headcount_cargo = (
        ativos
        .groupby(
            "Cargo"
        )
        .size()
        .sort_values(
            ascending=False
        )
    )

    if not headcount_cargo.empty:

        st.bar_chart(
            headcount_cargo
        )


    # =====================================================
    # CUSTO POR UNIDADE
    # =====================================================

    st.subheader(
        "🏢 Custo salarial por unidade"
    )

    folha_unidade = (
        ativos
        .groupby(
            "Unidade"
        )[
            "Salário"
        ]
        .sum()
        .sort_values(
            ascending=False
        )
    )

    if not folha_unidade.empty:

        st.bar_chart(
            folha_unidade
        )


    # =====================================================
    # RANKING DE SETORES
    # =====================================================

    st.subheader(
        "💰 Ranking de setores por folha salarial"
    )

    folha_setor = (
        ativos
        .groupby(
            "Setor"
        )[
            "Salário"
        ]
        .sum()
        .sort_values(
            ascending=False
        )
    )

    ranking_setores = (
        folha_setor
        .reset_index()
    )

    ranking_setores.columns = [
        "Setor",
        "Folha Salarial"
    ]

    ranking_setores[
        "Folha Salarial"
    ] = (
        ranking_setores[
            "Folha Salarial"
        ]
        .apply(
            moeda
        )
    )

    st.dataframe(
        ranking_setores,
        use_container_width=True,
        hide_index=True
    )


    # =====================================================
    # RESUMO POR SETOR
    # =====================================================

    resumo_setor_base = (
        ativos
        .groupby(
            "Setor",
            as_index=False
        )
        .agg(
            Headcount=(
                "Matrícula",
                "count"
            ),

            Folha_Salarial=(
                "Salário",
                "sum"
            ),

            Salario_Medio=(
                "Salário",
                "mean"
            )
        )
    )

    resumo_setor = (
        resumo_setor_base.copy()
    )

    resumo_setor[
        "Folha Salarial"
    ] = (
        resumo_setor[
            "Folha_Salarial"
        ]
        .apply(
            moeda
        )
    )

    resumo_setor[
        "Salário Médio"
    ] = (
        resumo_setor[
            "Salario_Medio"
        ]
        .apply(
            moeda
        )
    )

    resumo_setor = (
        resumo_setor[
            [
                "Setor",
                "Headcount",
                "Folha Salarial",
                "Salário Médio"
            ]
        ]
    )


    # =====================================================
    # RESUMO POR UNIDADE
    # =====================================================

    resumo_unidade_base = (
        ativos
        .groupby(
            "Unidade",
            as_index=False
        )
        .agg(
            Headcount=(
                "Matrícula",
                "count"
            ),

            Folha_Salarial=(
                "Salário",
                "sum"
            ),

            Salario_Medio=(
                "Salário",
                "mean"
            )
        )
    )

    resumo_unidade = (
        resumo_unidade_base.copy()
    )

    resumo_unidade[
        "Folha Salarial"
    ] = (
        resumo_unidade[
            "Folha_Salarial"
        ]
        .apply(
            moeda
        )
    )

    resumo_unidade[
        "Salário Médio"
    ] = (
        resumo_unidade[
            "Salario_Medio"
        ]
        .apply(
            moeda
        )
    )

    resumo_unidade = (
        resumo_unidade[
            [
                "Unidade",
                "Headcount",
                "Folha Salarial",
                "Salário Médio"
            ]
        ]
    )


    # =====================================================
    # TABELAS RESUMO
    # =====================================================

    st.subheader(
        "📋 Resumo por setor"
    )

    st.dataframe(
        resumo_setor,
        use_container_width=True,
        hide_index=True
    )


    st.subheader(
        "🏢 Resumo por unidade"
    )

    st.dataframe(
        resumo_unidade,
        use_container_width=True,
        hide_index=True
    )


    # =====================================================
    # MOVIMENTAÇÕES DETALHADAS
    # =====================================================

    st.subheader(
        f"🔄 Movimentações de {ano_analise}"
    )

    movimentacoes_detalhadas = (
        pd.concat(
            [
                admissoes_ano.assign(
                    Movimento="Admissão",
                    Data_Movimento=(
                        admissoes_ano[
                            "Admissão"
                        ]
                    )
                ),

                desligamentos_ano.assign(
                    Movimento="Desligamento",
                    Data_Movimento=(
                        desligamentos_ano[
                            "Demissão"
                        ]
                    )
                )
            ],
            ignore_index=True
        )
    )

    if not movimentacoes_detalhadas.empty:

        movimentacoes_detalhadas = (
            movimentacoes_detalhadas
            .sort_values(
                "Data_Movimento"
            )
        )

        st.dataframe(
            movimentacoes_detalhadas[
                [
                    "Data_Movimento",
                    "Movimento",
                    "Matrícula",
                    "Nome",
                    "Cargo",
                    "Setor",
                    "Unidade"
                ]
            ],
            use_container_width=True,
            hide_index=True
        )

    else:

        st.info(
            "Nenhuma movimentação encontrada."
        )


    # =====================================================
    # BASE DETALHADA
    # =====================================================

    st.subheader(
        "🗂️ Base detalhada"
    )

    base_exibicao = (
        dados_filtrados.copy()
    )

    base_exibicao.drop(
        columns=[
            "TempoEmpresaAnos"
        ],
        inplace=True
    )

    st.dataframe(
        base_exibicao,
        use_container_width=True,
        hide_index=True
    )


    # =====================================================
    # EXPORTAÇÃO
    # =====================================================

    st.subheader(
        "📥 Exportar análise"
    )

    indicadores = {
        "Ano analisado":
            ano_analise,

        "Headcount ativo":
            headcount,

        "Admissões":
            qtd_admissoes,

        "Desligamentos":
            qtd_desligamentos,

        "Folha salarial":
            moeda(folha),

        "Salário médio":
            moeda(salario_medio),

        "Turnover":
            percentual(turnover),

        "Tempo médio empresa":
            f"{tempo_medio:.1f} anos",

        "Maior salário":
            moeda(maior_salario),

        "Menor salário":
            moeda(menor_salario)
    }

    relatorio_excel = (
        gerar_relatorio_excel(
            base_exibicao,
            resumo_setor,
            resumo_unidade,
            movimentacoes,
            indicadores
        )
    )

    st.download_button(
        label="📊 Baixar relatório gerencial em Excel",
        data=relatorio_excel,
        file_name=(
            f"people_analytics_{ano_analise}.xlsx"
        ),
        mime=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )


    st.info(
        """
        Os indicadores são recalculados automaticamente
        conforme os filtros selecionados na barra lateral.
        """
    )